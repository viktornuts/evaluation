from __future__ import annotations

import json
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
IMPORT_DATASET_DIR = ROOT / "imports" / "datasets" / "customer_gold_v1"
XLSX_IMPORT_PATH = IMPORT_DATASET_DIR / "customer_gold_v1_test_cases.xlsx"
IMPORT_PATH = IMPORT_DATASET_DIR / "customer_gold_v1_test_cases.json"
DATASET_DIR = ROOT / "Документация" / "Датасеты" / "customer_gold_v1"
SOURCE_XLSX_PATH = DATASET_DIR / "test-cases" / "source" / "gold_test_cases.xlsx"
TEST_CASES_MD_PATH = DATASET_DIR / "test-cases" / "test-cases.md"

STEP_TYPE_PRECONDITION = "Предусловие"
STEP_TYPE_STEP = "Шаг"
STEP_TYPE_POSTCONDITION = "Постусловие"

DIRECTION_MAP = {
    "позитивный": "positive",
    "негативный": "negative",
}

TEST_CASE_TYPE_MAP = {
    "WEB": "WEB",
    "API": "API",
    "E2E": "E2E",
    "интеграционный": "INT",
}


def requirement_id(requirement_code: str) -> str:
    return f"req_customer_gold_{requirement_code.lower().replace('-', '_')}"


def test_case_id(test_case_code: str) -> str:
    return f"tc_customer_gold_{test_case_code.lower().replace('-', '_')}"


def parse_requirement_codes(value: object) -> list[str]:
    return re.findall(r"REQ-\d{2}", str(value or ""))


def parse_test_case_codes(value: object) -> list[str]:
    return re.findall(r"TC-\d{2}", str(value or ""))


def numbered_text(rows: list[dict[str, object]], key: str) -> str | None:
    parts = []
    for index, row in enumerate(rows, start=1):
        value = row.get(key)
        if value:
            parts.append(f"{index}. {value}")
    return "\n".join(parts) if parts else None


def parse_test_cases(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook.worksheets[0]

    parsed = []
    current: dict | None = None
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        tc_code, reqs, title, tc_type, direction, step_type, step_number, action, expected = row[:9]
        if tc_code:
            current = {
                "code": str(tc_code).strip(),
                "requirement_codes": parse_requirement_codes(reqs),
                "title": str(title).strip(),
                "test_case_type": TEST_CASE_TYPE_MAP.get(str(tc_type).strip(), str(tc_type).strip()),
                "direction": DIRECTION_MAP.get(str(direction).strip(), str(direction).strip()),
                "preconditions": [],
                "steps": [],
                "postconditions": [],
            }
            parsed.append(current)

        if current is None:
            continue

        row_data = {
            "step_number": int(step_number) if step_number is not None else None,
            "action": str(action).strip() if action is not None else None,
            "expected": str(expected).strip() if expected is not None else None,
        }
        if step_type == STEP_TYPE_PRECONDITION:
            current["preconditions"].append(row_data)
        elif step_type == STEP_TYPE_STEP:
            current["steps"].append(row_data)
        elif step_type == STEP_TYPE_POSTCONDITION:
            current["postconditions"].append(row_data)

    return parsed


def parse_coverage(workbook_path: Path) -> dict[str, list[str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook.worksheets[1]
    coverage: dict[str, list[str]] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        _, req_code, _, tc_codes, _ = row[:5]
        if not req_code:
            continue
        coverage[str(req_code).strip()] = parse_test_case_codes(tc_codes)
    return coverage


def build_payload(test_cases: list[dict], coverage: dict[str, list[str]]) -> dict:
    test_case_entries = []
    for tc in test_cases:
        tc_id = test_case_id(tc["code"])
        test_case_entries.append(
            {
                "id": tc_id,
                "test_case_code": tc["code"],
                "title": tc["title"],
                "test_case_type": tc["test_case_type"],
                "direction": tc["direction"],
                "origin": "expected",
                "preconditions": numbered_text(tc["preconditions"], "action"),
                "postconditions": numbered_text(tc["postconditions"], "expected"),
                "steps": [
                    {
                        "id": f"step_{tc_id}_{step['step_number']}",
                        "step_number": step["step_number"],
                        "action": step["action"] or "",
                        "expected_result": step["expected"] or "",
                        "linked_requirement_ids": [
                            requirement_id(req_code) for req_code in tc["requirement_codes"]
                        ],
                        "link_rationale": (
                            "Связь шага унаследована от связи тест-кейса с требованиями "
                            "в эталонном Excel; отдельной покроковой разметки в источнике нет."
                        ),
                        "linked_by": "customer_gold_v1_import",
                        "review_status": "approved",
                    }
                    for step in tc["steps"]
                ],
            }
        )

    links = []
    for req_code, tc_codes in coverage.items():
        for tc_code in tc_codes:
            tc = next(item for item in test_cases if item["code"] == tc_code)
            links.append(
                {
                    "id": (
                        f"rtcl_customer_gold_"
                        f"{req_code.lower().replace('-', '_')}_"
                        f"{tc_code.lower().replace('-', '_')}"
                    ),
                    "requirement_id": requirement_id(req_code),
                    "test_case_id": test_case_id(tc_code),
                    "coverage_status": "covered",
                    "coverage_type": tc["direction"],
                    "is_primary": True,
                    "rationale": "Связь импортирована из листа покрытия эталонных тест-кейсов.",
                    "linked_by": "customer_gold_v1_import",
                    "review_status": "approved",
                }
            )

    return {
        "dataset": {
            "id": "dataset_customer_gold_v1",
            "name": "customer_gold",
            "version": "v1",
            "description": (
                "Высокоприоритетный эталонный датасет от заказчика: "
                "требования и тест-кейсы по интеграции С.ФТ с С.Релизы."
            ),
            "status": "draft",
        },
        "cases": [
            {
                "id": "case_customer_gold_release_integration",
                "case_code": "GOLD-REQ-001",
                "title": "Интеграция С.ФТ с С.Релизы: эталонные требования и ТК",
                "description": (
                    "Эталонный набор требований и тест-кейсов из материалов заказчика."
                ),
                "case_type": "golden",
                "input_profile_code": "customer_gold_requirements",
                "input_profile_name": "Эталонные требования заказчика",
                "test_cases": test_case_entries,
                "requirement_test_case_links": links,
            }
        ],
    }


def validate(test_cases: list[dict], coverage: dict[str, list[str]]) -> None:
    tc_codes = {tc["code"] for tc in test_cases}
    coverage_tc_codes = {tc_code for tc_codes in coverage.values() for tc_code in tc_codes}
    missing_in_cases = sorted(coverage_tc_codes - tc_codes)
    missing_in_coverage = sorted(tc_codes - coverage_tc_codes)
    if missing_in_cases:
        raise ValueError(f"TC codes from coverage are absent in test cases sheet: {missing_in_cases}")
    if missing_in_coverage:
        raise ValueError(f"TC codes from test cases sheet are absent in coverage sheet: {missing_in_coverage}")

    req_codes = set(coverage)
    expected_req_codes = {f"REQ-{number:02d}" for number in range(1, 31)}
    if req_codes != expected_req_codes:
        raise ValueError(f"Coverage requirements mismatch: {sorted(expected_req_codes - req_codes)}")


def write_markdown(test_cases: list[dict], coverage: dict[str, list[str]]) -> None:
    TEST_CASES_MD_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Эталонные тест-кейсы customer_gold_v1",
        "",
        "Источник: `source/gold_test_cases.xlsx`.",
        "",
        "## Тест-кейсы",
        "",
        "| ТК | Требования | Название | Вид ТК | Направление | Шагов |",
        "|---|---|---|---|---|---:|",
    ]
    for tc in test_cases:
        lines.append(
            f"| {tc['code']} | {', '.join(tc['requirement_codes'])} | {tc['title']} | "
            f"{tc['test_case_type']} | {tc['direction']} | {len(tc['steps'])} |"
        )

    lines.extend(
        [
            "",
            "## Покрытие",
            "",
            "| Требование | Тест-кейсы |",
            "|---|---|",
        ]
    )
    for req_code in sorted(coverage):
        lines.append(f"| {req_code} | {', '.join(coverage[req_code])} |")

    TEST_CASES_MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    if not XLSX_IMPORT_PATH.exists():
        raise FileNotFoundError(f"Missing source workbook: {XLSX_IMPORT_PATH}")

    SOURCE_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(XLSX_IMPORT_PATH, SOURCE_XLSX_PATH)

    test_cases = parse_test_cases(XLSX_IMPORT_PATH)
    coverage = parse_coverage(XLSX_IMPORT_PATH)
    validate(test_cases, coverage)

    payload = build_payload(test_cases, coverage)
    IMPORT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown(test_cases, coverage)

    link_count = sum(len(tc_codes) for tc_codes in coverage.values())
    print(f"Wrote {IMPORT_PATH}")
    print(f"Test cases: {len(test_cases)}")
    print(f"Requirement coverage rows: {len(coverage)}")
    print(f"Requirement-test case links: {link_count}")


if __name__ == "__main__":
    main()
